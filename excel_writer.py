# excel_writer.py
"""Excel export functionality with enhanced format reporting."""

import logging
from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import Document

class ExcelWriter:
    """Handles writing document data to Excel files."""
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.workbook = Workbook()
        self.logger = logging.getLogger(__name__)
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def write_summary(self, documents: List[Document]):
        """Write summary table to Excel with enhanced format information."""
        # Create summary data
        summary_data = []
        for doc in documents:
            summary_data.append({
                'ID': doc.id,
                'Parent Folder': doc.parent_folder,
                'Document Name': doc.name,
                'Document Title': doc.title,
                'Author (Properties)': doc.author,
                'Author (Text)': getattr(doc, 'author_from_text', 'Unknown'),
                'Word Count': doc.word_count,
                'Image Count': doc.image_count,
                'Arabic References': getattr(doc, 'arabic_reference_count', 0),
                'English References': getattr(doc, 'english_reference_count', 0),
                'Footnotes': getattr(doc, 'footnote_count', 0),
                'Uses Proper Styles': 'Yes' if doc.uses_proper_styles else 'No',
                'Format Quality': getattr(doc, 'format_quality', 'Unknown'),
                'Total Headings': doc.heading_stats.get('total_headings', 0) if hasattr(doc, 'heading_stats') else 0,
                'Images Missing Captions': len(getattr(doc, 'images_missing_captions', [])),
                'Style Issues': len(getattr(doc, 'heading_hierarchy_issues', []))
            })
        
        # Create DataFrame
        df = pd.DataFrame(summary_data)
        
        # Create worksheet
        ws = self.workbook.create_sheet("Summary")
        
        # Write data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header(ws)
        
        # Apply conditional formatting for quality column
        quality_col = self._find_column_index(ws, 'Format Quality')
        if quality_col:
            for row in range(2, ws.max_row + 1):
                quality = ws.cell(row=row, column=quality_col).value
                if quality == 'Poor':
                    ws.cell(row=row, column=quality_col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                elif quality == 'Fair':
                    ws.cell(row=row, column=quality_col).fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
                elif quality == 'Good':
                    ws.cell(row=row, column=quality_col).fill = PatternFill(
                        start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
                    )
                elif quality == 'Excellent':
                    ws.cell(row=row, column=quality_col).fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
        
        # Highlight documents not using proper styles
        style_col = self._find_column_index(ws, 'Uses Proper Styles')
        if style_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=style_col).value == 'No':
                    ws.cell(row=row, column=style_col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
        
        # Highlight documents with missing captions
        caption_col = self._find_column_index(ws, 'Images Missing Captions')
        if caption_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=caption_col).value > 0:
                    ws.cell(row=row, column=caption_col).fill = PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        self.logger.info(f"Written summary for {len(documents)} documents")
        
        # Add format issues sheet if there are any issues
        if any(getattr(doc, 'total_format_issues', 0) > 0 for doc in documents):
            self._write_format_issues_sheet(documents)
    
    def _write_format_issues_sheet(self, documents: List[Document]):
        """Create a sheet listing all format issues found."""
        ws = self.workbook.create_sheet("Format Issues")
        
        # Add headers
        ws.append(['Document ID', 'Document Name', 'Issue Type', 'Description', 'Suggested Action'])
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add issues
        for doc in documents:
            # Add heading hierarchy issues
            for issue in getattr(doc, 'heading_hierarchy_issues', []):
                ws.append([
                    doc.id,
                    doc.name,
                    'Heading Style',
                    f"{issue['heading']} - Current: {issue['current_style']}",
                    f"Change to: {issue['suggested_style']}"
                ])
            
            # Add images missing captions
            for image in getattr(doc, 'images_missing_captions', []):
                ws.append([
                    doc.id,
                    doc.name,
                    'Missing Caption',
                    image,
                    'Add caption below image'
                ])
            
            # Add general format issues
            for issue in getattr(doc, 'format_issues', []):
                ws.append([
                    doc.id,
                    doc.name,
                    'Format Issue',
                    issue,
                    'Apply proper Word styles'
                ])
        
        # Auto-adjust columns
        self._adjust_column_widths(ws)
        
        self.logger.info(f"Written format issues sheet")
    
    def write_sections(self, documents: List[Document]):
        """Write section details for each document (only headings and images)."""
        for doc in documents:
            if not doc.sections:
                continue
            
            # Create sheet name with only document ID
            sheet_name = f"Doc_{doc.id}"
            
            # Create worksheet
            ws = self.workbook.create_sheet(sheet_name)
            
            # Add document info
            ws.append(['Document ID:', doc.id])
            ws.append(['Document Name:', doc.name])
            ws.append(['Document Title:', doc.title])
            ws.append(['Author (Properties):', doc.author])
            ws.append(['Author (Text):', getattr(doc, 'author_from_text', 'Unknown')])
            ws.append(['Format Quality:', getattr(doc, 'format_quality', 'Unknown')])
            ws.append(['Uses Proper Styles:', 'Yes' if doc.uses_proper_styles else 'No'])
            ws.append(['Arabic References:', getattr(doc, 'arabic_reference_count', 0)])
            ws.append(['English References:', getattr(doc, 'english_reference_count', 0)])
            ws.append(['Footnotes:', getattr(doc, 'footnote_count', 0)])
            ws.append([])  # Empty row
            
            # Add headers for sections
            ws.append(['Type', 'Content', 'Current Style', 'Suggested Style', 'Font', 'Size', 'Issue'])
            
            # Format section header row
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add only headings and images
            for section in doc.sections:
                if section.section_type in ['heading', 'image']:
                    # Determine if there's an issue
                    issue = ""
                    if section.section_type == 'heading':
                        if section.style_name != section.suggested_style and section.suggested_style != "Unknown":
                            issue = f"Should be {section.suggested_style}"
                    elif section.section_type == 'image':
                        if not section.has_caption:
                            issue = "Missing caption"
                    
                    ws.append([
                        section.section_type.capitalize(),
                        section.heading[:100] + "..." if len(section.heading) > 100 else section.heading,
                        section.style_name or "Normal",
                        section.suggested_style or "N/A",
                        section.font_name or "N/A",
                        section.font_size or "N/A",
                        issue
                    ])
                    
                    # Highlight rows with issues
                    if issue:
                        for col in range(1, 8):
                            ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                            )
                    
                    # Highlight if using Normal style for headings
                    if section.section_type == 'heading' and section.style_name == "Normal":
                        ws.cell(row=ws.max_row, column=3).font = Font(color="FF0000", bold=True)
            
            # Format document info cells
            for row in range(1, 11):  # Updated to include new rows
                ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Highlight format quality
            if getattr(doc, 'format_quality', '') in ['Poor', 'Fair']:
                ws.cell(row=6, column=2).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            self.logger.info(f"Written sections for document {doc.id}")
    
    def save(self):
        """Save the workbook to file."""
        self.workbook.save(self.output_path)
        self.logger.info(f"Excel file saved to: {self.output_path}")
    
    def _format_header(self, worksheet):
        """Format header row with bold font and background color."""
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility."""
        # Remove invalid characters
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Limit to 31 characters
        return name[:31]
    
    def _find_column_index(self, worksheet, column_name: str) -> Optional[int]:
        """Find column index by header name."""
        for idx, cell in enumerate(worksheet[1], 1):
            if cell.value == column_name:
                return idx
        return None